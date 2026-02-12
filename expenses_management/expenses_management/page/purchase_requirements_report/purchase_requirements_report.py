import frappe
from frappe import _
from frappe.utils import today, getdate, flt
from collections import defaultdict
import json
import io


def has_permission_to_view():
	allowed_roles = [
		"System Manager", "Accounts Manager", "Accounts User",
		"Sales Manager", "Sales User", "Stock Manager", "Stock User",
		"Purchase Manager", "Purchase User"
	]
	user_roles = frappe.get_roles()
	for role in allowed_roles:
		if role in user_roles:
			return True
	return False


@frappe.whitelist()
def get_filter_options():
	if not has_permission_to_view():
		frappe.throw(_("You don't have permission to view this report"))

	companies = frappe.db.sql("SELECT name FROM `tabCompany` ORDER BY name", as_list=1)
	companies = [c[0] for c in companies]

	item_groups = frappe.db.sql("""
		SELECT name FROM `tabItem Group`
		WHERE parent_item_group = 'Beam'
		ORDER BY name
	""", as_list=1)
	item_groups = [g[0] for g in item_groups]

	lengths = []
	if item_groups:
		lengths = frappe.db.sql("""
			SELECT DISTINCT custom_length
			FROM `tabItem`
			WHERE item_group IN %(groups)s
			AND custom_length IS NOT NULL AND custom_length > 0
			AND disabled = 0 AND is_stock_item = 1
			ORDER BY custom_length
		""", {"groups": tuple(item_groups)}, as_list=1)
		lengths = [float(l[0]) for l in lengths]

	return {
		"companies": companies,
		"item_groups": item_groups,
		"lengths": lengths
	}


@frappe.whitelist()
def get_warehouses_for_company(company):
	if not has_permission_to_view():
		frappe.throw(_("You don't have permission to view this report"))

	warehouses = frappe.db.sql("""
		SELECT name FROM `tabWarehouse`
		WHERE company = %(company)s
		AND is_group = 0 AND disabled = 0
		ORDER BY name
	""", {"company": company}, as_list=1)
	return [w[0] for w in warehouses]


@frappe.whitelist()
def get_purchase_requirements_data(company, from_date, to_date,
	item_groups=None, warehouses=None, lengths=None):

	if not has_permission_to_view():
		frappe.throw(_("You don't have permission to view this report"))

	if not company:
		frappe.throw(_("Company is required"))

	from_date = getdate(from_date) if from_date else getdate(today())
	to_date = getdate(to_date) if to_date else getdate(today())

	if isinstance(item_groups, str):
		item_groups = json.loads(item_groups)
	if isinstance(warehouses, str):
		warehouses = json.loads(warehouses)
	if isinstance(lengths, str):
		lengths = json.loads(lengths)

	if not item_groups:
		item_groups = frappe.db.sql("""
			SELECT name FROM `tabItem Group`
			WHERE parent_item_group = 'Beam'
		""", as_list=1)
		item_groups = [g[0] for g in item_groups]

	if not warehouses:
		warehouses = frappe.db.sql("""
			SELECT name FROM `tabWarehouse`
			WHERE company = %(company)s AND is_group = 0 AND disabled = 0
			ORDER BY name
		""", {"company": company}, as_list=1)
		warehouses = [w[0] for w in warehouses]

	if not item_groups or not warehouses:
		return {
			"items": [],
			"warehouses": warehouses or [],
			"filters": {
				"company": company,
				"from_date": str(from_date),
				"to_date": str(to_date),
				"item_groups": item_groups or [],
				"lengths": lengths or []
			}
		}

	item_conditions = "WHERE item.item_group IN %(item_groups)s AND item.disabled = 0 AND item.is_stock_item = 1"
	item_values = {"item_groups": tuple(item_groups)}

	if lengths:
		float_lengths = [float(l) for l in lengths]
		item_conditions += " AND item.custom_length IN %(lengths)s"
		item_values["lengths"] = tuple(float_lengths)

	items = frappe.db.sql("""
		SELECT
			item.item_code,
			item.item_name,
			COALESCE(item.weight_per_unit, 0) as weight_per_unit,
			COALESCE(item.custom_length, 0) as custom_length,
			item.item_group
		FROM `tabItem` item
		{conditions}
		ORDER BY item.item_group, item.custom_length, item.item_name
	""".format(conditions=item_conditions), item_values, as_dict=1)

	if not items:
		return {
			"items": [],
			"warehouses": warehouses,
			"filters": {
				"company": company,
				"from_date": str(from_date),
				"to_date": str(to_date),
				"item_groups": item_groups,
				"lengths": lengths or []
			}
		}

	item_codes = tuple([i.item_code for i in items])

	bin_data = frappe.db.sql("""
		SELECT
			b.item_code,
			b.warehouse,
			COALESCE(b.actual_qty, 0) as actual_qty,
			COALESCE(b.ordered_qty, 0) as ordered_qty
		FROM `tabBin` b
		WHERE b.item_code IN %(item_codes)s
		AND b.warehouse IN %(warehouses)s
	""", {
		"item_codes": item_codes,
		"warehouses": tuple(warehouses)
	}, as_dict=1)

	stock_map = defaultdict(lambda: defaultdict(lambda: {"actual_qty": 0, "ordered_qty": 0}))
	for b in bin_data:
		stock_map[b.item_code][b.warehouse]["actual_qty"] = flt(b.actual_qty)
		stock_map[b.item_code][b.warehouse]["ordered_qty"] = flt(b.ordered_qty)

	sales_data = frappe.db.sql("""
		SELECT
			sii.item_code,
			sii.warehouse,
			SUM(sii.stock_qty) as total_sold_qty
		FROM `tabSales Invoice Item` sii
		INNER JOIN `tabSales Invoice` si ON si.name = sii.parent
		WHERE si.docstatus = 1
		AND si.company = %(company)s
		AND si.is_return = 0
		AND si.posting_date BETWEEN %(from_date)s AND %(to_date)s
		AND sii.item_code IN %(item_codes)s
		AND sii.warehouse IN %(warehouses)s
		GROUP BY sii.item_code, sii.warehouse
	""", {
		"company": company,
		"from_date": from_date,
		"to_date": to_date,
		"item_codes": item_codes,
		"warehouses": tuple(warehouses)
	}, as_dict=1)

	sales_map = defaultdict(lambda: defaultdict(float))
	for s in sales_data:
		if s.warehouse:
			sales_map[s.item_code][s.warehouse] = flt(s.total_sold_qty)

	result_items = []
	for item in items:
		ic = item.item_code
		item_stock = {}
		item_sales = {}

		for wh in warehouses:
			sd = stock_map[ic][wh]
			item_stock[wh] = {
				"actual_qty": sd["actual_qty"],
				"ordered_qty": sd["ordered_qty"]
			}
			item_sales[wh] = sales_map[ic].get(wh, 0)

		result_items.append({
			"item_code": ic,
			"item_name": item.item_name,
			"weight_per_unit": item.weight_per_unit,
			"custom_length": item.custom_length,
			"item_group": item.item_group,
			"stock_data": item_stock,
			"sales_data": item_sales
		})

	return {
		"items": result_items,
		"warehouses": warehouses,
		"filters": {
			"company": company,
			"from_date": str(from_date),
			"to_date": str(to_date),
			"item_groups": item_groups,
			"lengths": lengths or []
		}
	}


@frappe.whitelist()
def export_excel(company, from_date, to_date,
	item_groups=None, warehouses=None, lengths=None, actual_required=None):

	from openpyxl import Workbook
	from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
	from openpyxl.utils import get_column_letter

	if not has_permission_to_view():
		frappe.throw(_("You don't have permission to view this report"))

	# Get same data as the report
	result = get_purchase_requirements_data(company, from_date, to_date,
		item_groups, warehouses, lengths)

	items = result.get("items", [])
	wh_list = result.get("warehouses", [])
	filters = result.get("filters", {})

	if isinstance(actual_required, str):
		actual_required = json.loads(actual_required)
	if not actual_required:
		actual_required = {}

	wb = Workbook()
	ws = wb.active
	ws.title = "شيت المشتريات"
	ws.sheet_view.rightToLeft = True

	# Styles
	header_font = Font(name='Segoe UI', bold=True, size=12, color='FFFFFF')
	sub_header_font = Font(name='Segoe UI', bold=True, size=10, color='FFFFFF')
	data_font = Font(name='Segoe UI', bold=True, size=10)
	total_font = Font(name='Segoe UI', bold=True, size=11, color='1e293b')

	dark_fill = PatternFill(start_color='0f172a', end_color='1e293b', fill_type='solid')
	item_fill = PatternFill(start_color='eef2ff', end_color='eef2ff', fill_type='solid')
	item_fill_alt = PatternFill(start_color='e0e7ff', end_color='e0e7ff', fill_type='solid')
	wh_even_fill = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
	wh_odd_fill = PatternFill(start_color='2d1b69', end_color='2d1b69', fill_type='solid')
	wh_sub_even = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
	wh_sub_odd = PatternFill(start_color='2d1b69', end_color='2d1b69', fill_type='solid')
	data_even_fill = PatternFill(start_color='f8fafc', end_color='f8fafc', fill_type='solid')
	data_odd_fill = PatternFill(start_color='f0f4ff', end_color='f0f4ff', fill_type='solid')
	total_fill = PatternFill(start_color='f1f5f9', end_color='f1f5f9', fill_type='solid')

	thin_border = Border(
		left=Side(style='thin', color='cbd5e1'),
		right=Side(style='thin', color='cbd5e1'),
		top=Side(style='thin', color='cbd5e1'),
		bottom=Side(style='thin', color='cbd5e1')
	)
	center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
	right_align = Alignment(horizontal='right', vertical='center', wrap_text=True)

	# Column layout:
	# 6 item detail cols + 7 per warehouse + 5 total cols
	item_cols = ['#', 'رقم الصنف', 'اسم الصنف', 'الوزن', 'الطول', 'نوع الجسر']
	wh_sub_cols = ['المخزون', 'المتبقى', 'الإجمالى', 'المبيعات', 'المطلوب حبة', 'المطلوب طن', 'المطلوب الفعلى']
	total_cols = ['اجمالى المخزون حبة', 'اجمالى المخزون طن', 'اجمالى المبيعات حبة', 'اجمالى المبيعات طن', 'اجمالى المطلوب الفعلى']

	num_item_cols = len(item_cols)
	num_wh_subs = len(wh_sub_cols)
	num_total_cols = len(total_cols)
	total_columns = num_item_cols + (num_wh_subs * len(wh_list)) + num_total_cols

	# ---- Row 1: Group headers with merged cells ----
	row1 = 1
	# Item details merged header
	ws.merge_cells(start_row=row1, start_column=1, end_row=row1, end_column=num_item_cols)
	cell = ws.cell(row=row1, column=1, value='بيانات الصنف')
	cell.font = header_font
	cell.fill = dark_fill
	cell.alignment = center_align
	cell.border = thin_border

	# Warehouse merged headers
	for wh_idx, wh in enumerate(wh_list):
		wh_short = wh.replace(' - م', '')
		start_col = num_item_cols + 1 + (wh_idx * num_wh_subs)
		end_col = start_col + num_wh_subs - 1
		ws.merge_cells(start_row=row1, start_column=start_col, end_row=row1, end_column=end_col)
		cell = ws.cell(row=row1, column=start_col, value=wh_short)
		cell.font = header_font
		cell.fill = wh_even_fill if wh_idx % 2 == 0 else wh_odd_fill
		cell.alignment = center_align
		cell.border = thin_border

	# Totals merged header
	totals_start = num_item_cols + 1 + (len(wh_list) * num_wh_subs)
	ws.merge_cells(start_row=row1, start_column=totals_start, end_row=row1, end_column=total_columns)
	cell = ws.cell(row=row1, column=totals_start, value='الإجماليات')
	cell.font = header_font
	cell.fill = dark_fill
	cell.alignment = center_align
	cell.border = thin_border

	# Fill borders for merged cells
	for col in range(1, total_columns + 1):
		c = ws.cell(row=row1, column=col)
		c.border = thin_border
		if not c.fill or c.fill.start_color.rgb == '00000000':
			c.fill = dark_fill
			c.font = header_font
			c.alignment = center_align

	# ---- Row 2: Sub-headers ----
	row2 = 2
	for i, name in enumerate(item_cols):
		cell = ws.cell(row=row2, column=i + 1, value=name)
		cell.font = sub_header_font
		cell.fill = dark_fill
		cell.alignment = center_align
		cell.border = thin_border

	for wh_idx in range(len(wh_list)):
		for sub_idx, sub_name in enumerate(wh_sub_cols):
			col = num_item_cols + 1 + (wh_idx * num_wh_subs) + sub_idx
			cell = ws.cell(row=row2, column=col, value=sub_name)
			cell.font = sub_header_font
			cell.fill = wh_sub_even if wh_idx % 2 == 0 else wh_sub_odd
			cell.alignment = center_align
			cell.border = thin_border

	for i, name in enumerate(total_cols):
		col = totals_start + i
		cell = ws.cell(row=row2, column=col, value=name)
		cell.font = sub_header_font
		cell.fill = dark_fill
		cell.alignment = center_align
		cell.border = thin_border

	# Style for group separator rows
	group_font = Font(name='Segoe UI', bold=True, size=12, color='FFFFFF')
	group_fill = PatternFill(start_color='6366f1', end_color='8b5cf6', fill_type='solid')

	# ---- Data rows ----
	data_row = 3
	visible_idx = 0
	current_group = ''
	for item in items:
		# Skip all-zero rows
		has_any = False
		for wh in wh_list:
			sd = item.get("stock_data", {}).get(wh, {})
			sl = item.get("sales_data", {}).get(wh, 0)
			if flt(sd.get("actual_qty", 0)) != 0 or flt(sd.get("ordered_qty", 0)) != 0 or flt(sl) != 0:
				has_any = True
				break
		if not has_any:
			continue

		# Add group separator row when item_group changes
		if item.get("item_group") != current_group:
			current_group = item.get("item_group", "")
			ws.merge_cells(start_row=data_row, start_column=1, end_row=data_row, end_column=total_columns)
			cell = ws.cell(row=data_row, column=1, value=current_group)
			cell.font = group_font
			cell.fill = group_fill
			cell.alignment = Alignment(horizontal='right', vertical='center')
			cell.border = thin_border
			for c in range(2, total_columns + 1):
				ws.cell(row=data_row, column=c).border = thin_border
				ws.cell(row=data_row, column=c).fill = group_fill
			ws.row_dimensions[data_row].height = 28
			data_row += 1

		row_fill = item_fill if visible_idx % 2 == 0 else item_fill_alt
		visible_idx += 1

		# Item details
		item_values = [visible_idx, item["item_code"], item["item_name"], item["weight_per_unit"], item["custom_length"], item["item_group"]]
		for i, val in enumerate(item_values):
			cell = ws.cell(row=data_row, column=i + 1, value=val)
			cell.font = data_font
			cell.fill = row_fill
			cell.alignment = right_align if i == 2 else center_align
			cell.border = thin_border

		grand_stock = 0
		grand_ordered = 0
		grand_sales = 0

		for wh_idx, wh in enumerate(wh_list):
			wh_fill = data_even_fill if wh_idx % 2 == 0 else data_odd_fill
			sd = item.get("stock_data", {}).get(wh, {})
			actual_qty = flt(sd.get("actual_qty", 0))
			ordered_qty = flt(sd.get("ordered_qty", 0))
			total_qty = actual_qty + ordered_qty
			sales_qty = flt(item.get("sales_data", {}).get(wh, 0))
			required_pcs = max(0, (sales_qty - total_qty)) * 1.1
			required_tons = required_pcs * flt(item.get("weight_per_unit", 0)) / 1000

			grand_stock += actual_qty
			grand_ordered += ordered_qty
			grand_sales += sales_qty

			# Get actual required from input data
			ar_key = "{}|{}".format(item["item_code"], wh)
			actual_req = flt(actual_required.get(ar_key, 0))

			wh_values = [actual_qty, ordered_qty, total_qty, sales_qty, round(required_pcs), round(required_tons, 2), actual_req or '']
			for sub_idx, val in enumerate(wh_values):
				col = num_item_cols + 1 + (wh_idx * num_wh_subs) + sub_idx
				cell = ws.cell(row=data_row, column=col, value=val if val != '' else None)
				cell.font = data_font
				cell.fill = wh_fill
				cell.alignment = center_align
				cell.border = thin_border

		grand_total = grand_stock + grand_ordered
		grand_stock_tons = grand_total * flt(item.get("weight_per_unit", 0)) / 1000
		grand_sales_tons = grand_sales * flt(item.get("weight_per_unit", 0)) / 1000

		# Sum actual required for this item
		grand_actual_req = 0
		for wh in wh_list:
			ar_key = "{}|{}".format(item["item_code"], wh)
			grand_actual_req += flt(actual_required.get(ar_key, 0))

		total_values = [grand_total, round(grand_stock_tons, 2), grand_sales, round(grand_sales_tons, 2), grand_actual_req or '']
		for i, val in enumerate(total_values):
			col = totals_start + i
			cell = ws.cell(row=data_row, column=col, value=val if val != '' else None)
			cell.font = total_font
			cell.fill = total_fill
			cell.alignment = center_align
			cell.border = thin_border

		data_row += 1

	# Auto-fit column widths
	for col in range(1, total_columns + 1):
		col_letter = get_column_letter(col)
		if col == 1:  # index #
			ws.column_dimensions[col_letter].width = 5
		elif col == 2:  # item code
			ws.column_dimensions[col_letter].width = 14
		elif col == 3:  # item name - wider
			ws.column_dimensions[col_letter].width = 28
		else:
			ws.column_dimensions[col_letter].width = 12

	# Freeze panes: freeze first 2 rows and first 5 columns
	ws.freeze_panes = 'A3'

	# Row heights
	ws.row_dimensions[1].height = 30
	ws.row_dimensions[2].height = 25

	# Save to BytesIO
	output = io.BytesIO()
	wb.save(output)
	output.seek(0)
	xlsx_data = output.getvalue()

	filename = "purchase_requirements_{}_{}.xlsx".format(filters.get("from_date", ""), filters.get("to_date", ""))

	frappe.local.response.filename = filename
	frappe.local.response.filecontent = xlsx_data
	frappe.local.response.type = "download"
	frappe.local.response.content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
